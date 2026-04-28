"""Flask blueprint for the feedback / benchmark / regression UI.

Mounted under /feedback. Self-contained: pulls config/path from
``current_app.config`` (set by ``create_app``). Does not modify any
existing webapp routes.
"""
from __future__ import annotations

import io
import json
import logging
import threading
import uuid
from datetime import datetime
from pathlib import Path
from typing import Optional

import yaml
from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)

from src.feedback import (
    AGREEMENT_TYPES,
    LINE_VERDICTS,
    OVERALL_VERDICTS,
    FeedbackEntry,
    FeedbackStore,
    LineFeedback,
    MissedDiscrepancy,
    build_entry_from_pred,
)
from src.feedback_insights import (
    find_miss_patterns,
    overall,
    per_agreement_type,
    per_config,
    per_supplier,
)
from src.feedback_loader import list_runs, load_run
from src.golden_set import compare, jobs_for_regression
from src.auto_tune import (
    AutoTuneStore,
    analyze_with_llm,
    apply_suggestion,
)
from src.prompt_history import PromptHistory, PromptVersion

log = logging.getLogger(__name__)

bp = Blueprint("feedback", __name__, url_prefix="/feedback")

# Regression runs are kept in-memory similarly to the main RUNS dict.
_REGRESSION_RUNS: dict[str, dict] = {}
_REGRESSION_LOCK = threading.Lock()


def _cfg() -> dict:
    cfg_path = Path(current_app.config["CFG_PATH"])
    c = yaml.safe_load(cfg_path.read_text(encoding="utf-8"))
    base = Path(current_app.config["BASE_DIR"])
    if isinstance(c.get("paths"), dict):
        for k, v in list(c["paths"].items()):
            if not isinstance(v, str):
                continue
            p = Path(v)
            if not p.is_absolute():
                c["paths"][k] = str((base / p).resolve())
    return c


def _store() -> FeedbackStore:
    base = Path(current_app.config["BASE_DIR"])
    path = base / "data" / "feedback.jsonl"
    return FeedbackStore(path)


def _autotune_store() -> AutoTuneStore:
    base = Path(current_app.config["BASE_DIR"])
    return AutoTuneStore(base / "data" / "autotune.json")


def _prompt_history() -> PromptHistory:
    base = Path(current_app.config["BASE_DIR"])
    return PromptHistory(base / "data" / "prompt_history.jsonl")


def _prompts_path() -> Path:
    base = Path(current_app.config["BASE_DIR"])
    cfg = _cfg()
    return Path(
        cfg.get("paths", {}).get(
            "prompts_file", str(base / "data" / "prompts.json")
        )
    )


def _results_dir() -> Path:
    return Path(_cfg()["paths"]["results_dir"])


@bp.route("/")
def index():
    """Dashboard: overview + recent feedback + recent runs available for review."""
    store = _store()
    entries = store.list_all()
    summary = overall(entries)
    runs = list_runs(_results_dir())[:10]
    recent = sorted(entries, key=lambda e: e.timestamp, reverse=True)[:10]
    return render_template(
        "feedback_index.html",
        summary=summary,
        recent_feedback=recent,
        runs=runs,
        total_entries=len(entries),
    )


@bp.route("/run/<path:filename>")
def pick_pair(filename: str):
    """Show the matrix of (invoice, config) cells from a past run, with status
    indicating whether each already has feedback."""
    xlsx = _results_dir() / filename
    if not xlsx.exists():
        abort(404)
    pairs = load_run(xlsx)
    store = _store()
    enriched = []
    for p in pairs:
        existing = store.find(p.invoice_id, p.config_name)
        enriched.append(
            {
                "invoice_id": p.invoice_id,
                "config_name": p.config_name,
                "supplier": (p.model_pred or {}).get("supplier_name") or "—",
                "invoice_number": (p.model_pred or {}).get("invoice_number") or "—",
                "discrepancy_count": sum(
                    1
                    for li in (p.model_pred or {}).get("line_items") or []
                    if li.get("has_discrepancy")
                ),
                "has_feedback": existing is not None,
                "feedback_id": existing.feedback_id if existing else None,
                "verdict": existing.overall_verdict if existing else None,
            }
        )
    return render_template(
        "feedback_pick.html", filename=filename, pairs=enriched
    )


@bp.route("/review/<path:filename>/<invoice_id>/<config_name>", methods=["GET"])
def review(filename: str, invoice_id: str, config_name: str):
    """Render the review form. Pre-populates from existing feedback if any."""
    xlsx = _results_dir() / filename
    if not xlsx.exists():
        abort(404)
    pairs = load_run(xlsx)
    target: Optional = None
    for p in pairs:
        if p.invoice_id == invoice_id and p.config_name == config_name:
            target = p
            break
    if target is None:
        abort(404)

    store = _store()
    existing = store.find(invoice_id, config_name)
    if existing is None:
        existing = build_entry_from_pred(
            invoice_id=invoice_id,
            config_name=config_name,
            pred=target.model_pred,
            run_filename=filename,
        )
    line_items = (target.model_pred or {}).get("line_items") or []
    return render_template(
        "feedback_review.html",
        filename=filename,
        entry=existing,
        line_items=line_items,
        agreement_types=AGREEMENT_TYPES,
        line_verdicts=LINE_VERDICTS,
        overall_verdicts=OVERALL_VERDICTS,
        is_existing=store.find(invoice_id, config_name) is not None,
    )


@bp.route("/review/<path:filename>/<invoice_id>/<config_name>", methods=["POST"])
def save_review(filename: str, invoice_id: str, config_name: str):
    store = _store()
    existing = store.find(invoice_id, config_name)
    feedback_id = existing.feedback_id if existing else uuid.uuid4().hex
    timestamp = datetime.now().isoformat(timespec="seconds")

    # Per-line verdicts. Form fields are named:
    #   line_verdict_<line_index>, line_correct_unit_<i>, line_correct_agreed_<i>,
    #   line_correct_diff_<i>, line_notes_<i>
    line_feedbacks: list[LineFeedback] = []
    indices = sorted({
        int(k.split("_")[-1])
        for k in request.form.keys()
        if k.startswith("line_verdict_")
    })
    for i in indices:
        verdict = request.form.get(f"line_verdict_{i}", "correct")
        if verdict not in LINE_VERDICTS:
            verdict = "correct"
        # Snapshot of model claim — re-loaded so we don't trust client state for the snapshot.
        xlsx = _results_dir() / filename
        pairs = load_run(xlsx) if xlsx.exists() else []
        snapshot = {}
        for p in pairs:
            if p.invoice_id == invoice_id and p.config_name == config_name:
                lis = (p.model_pred or {}).get("line_items") or []
                if 0 <= i < len(lis):
                    snapshot = lis[i]
                break
        line_feedbacks.append(
            LineFeedback(
                line_index=i,
                model_item_number=(
                    str(snapshot.get("item_number")).strip()
                    if snapshot.get("item_number") is not None
                    else None
                ),
                model_description=(snapshot.get("description") or "").strip() or None,
                model_quantity=_to_float(snapshot.get("quantity")),
                model_unit_price=_to_float(snapshot.get("unit_price")),
                model_agreed_unit_price=_to_float(snapshot.get("agreed_unit_price")),
                model_has_discrepancy=bool(snapshot.get("has_discrepancy")),
                model_discrepancy_amount=_to_float(snapshot.get("discrepancy_amount")),
                verdict=verdict,
                correct_unit_price=_to_float(request.form.get(f"line_correct_unit_{i}")),
                correct_agreed_unit_price=_to_float(
                    request.form.get(f"line_correct_agreed_{i}")
                ),
                correct_discrepancy_amount=_to_float(
                    request.form.get(f"line_correct_diff_{i}")
                ),
                notes=(request.form.get(f"line_notes_{i}") or "").strip() or None,
            )
        )

    # Missed discrepancies. Form fields:
    #   missed_item_<n>, missed_desc_<n>, missed_qty_<n>, missed_unit_<n>,
    #   missed_agreed_<n>, missed_diff_<n>, missed_notes_<n>
    missed: list[MissedDiscrepancy] = []
    missed_indices = sorted({
        int(k.split("_")[-1])
        for k in request.form.keys()
        if k.startswith("missed_desc_")
    })
    for n in missed_indices:
        desc = (request.form.get(f"missed_desc_{n}") or "").strip()
        if not desc:
            # Empty rows from the form's "blank starter"; skip.
            continue
        missed.append(
            MissedDiscrepancy(
                item_number=(request.form.get(f"missed_item_{n}") or "").strip() or None,
                description=desc,
                quantity=_to_float(request.form.get(f"missed_qty_{n}")),
                unit_price=_to_float(request.form.get(f"missed_unit_{n}")),
                correct_agreed_unit_price=_to_float(
                    request.form.get(f"missed_agreed_{n}")
                ),
                discrepancy_amount=_to_float(request.form.get(f"missed_diff_{n}")),
                notes=(request.form.get(f"missed_notes_{n}") or "").strip() or None,
            )
        )

    entry = FeedbackEntry(
        feedback_id=feedback_id,
        timestamp=timestamp,
        invoice_id=invoice_id,
        config_name=config_name,
        run_filename=filename,
        supplier_extracted=(request.form.get("supplier_extracted") or "").strip() or None,
        invoice_number_extracted=(
            request.form.get("invoice_number_extracted") or ""
        ).strip()
        or None,
        supplier_canonical=(request.form.get("supplier_canonical") or "").strip() or None,
        agreement_type=(
            request.form.get("agreement_type") or "price_agreement"
        ),
        agreement_notes=(request.form.get("agreement_notes") or "").strip() or None,
        line_feedbacks=line_feedbacks,
        missed_discrepancies=missed,
        overall_verdict=request.form.get("overall_verdict") or "acceptable",
        overall_notes=(request.form.get("overall_notes") or "").strip() or None,
    )
    store.save(entry)
    flash(
        f"Feedback saved: {len(line_feedbacks)} line verdicts, "
        f"{len(missed)} missed discrepancies."
    )
    return redirect(url_for("feedback.pick_pair", filename=filename))


@bp.route("/delete/<feedback_id>", methods=["POST"])
def delete_feedback(feedback_id: str):
    store = _store()
    if store.delete(feedback_id):
        flash("Feedback deleted.")
    else:
        flash("Feedback not found.")
    return redirect(url_for("feedback.index"))


@bp.route("/insights")
def insights():
    store = _store()
    entries = store.list_all()
    return render_template(
        "feedback_insights.html",
        summary=overall(entries),
        per_config=per_config(entries),
        per_supplier=per_supplier(entries),
        per_agreement_type=per_agreement_type(entries),
        miss_patterns=find_miss_patterns(entries),
        total=len(entries),
    )


@bp.route("/regression")
def regression_index():
    """Show what's available to re-run + history of past regression runs."""
    store = _store()
    entries = store.list_all()
    invoice_ids = sorted({e.invoice_id for e in entries})
    config_names = sorted({e.config_name for e in entries})
    cfg = _cfg()
    invoices_dir = Path(cfg["paths"]["invoices_dir"])
    on_disk = {p.name for p in invoices_dir.iterdir()} if invoices_dir.exists() else set()
    available = [iid for iid in invoice_ids if iid in on_disk]
    missing = [iid for iid in invoice_ids if iid not in on_disk]
    with _REGRESSION_LOCK:
        runs = list(_REGRESSION_RUNS.values())
    runs.sort(key=lambda r: r.get("started_at", ""), reverse=True)
    return render_template(
        "feedback_regression.html",
        feedback_count=len(entries),
        invoice_ids=invoice_ids,
        config_names=config_names,
        available_count=len(available),
        missing_invoices=missing,
        runs=runs[:10],
    )


@bp.route("/regression/start", methods=["POST"])
def regression_start():
    """Re-run feedback'd invoices through current prompts and compare."""
    store = _store()
    entries = store.list_all()
    if not entries:
        flash("No feedback yet — nothing to regress against.")
        return redirect(url_for("feedback.regression_index"))

    cfg = _cfg()
    invoices_dir = Path(cfg["paths"]["invoices_dir"])
    agreements_dir = Path(cfg["paths"]["agreements_dir"])
    jobs, fb_index = jobs_for_regression(entries, invoices_dir, agreements_dir)
    if not jobs:
        flash(
            "None of the invoices in feedback are currently on disk. Re-upload them first."
        )
        return redirect(url_for("feedback.regression_index"))

    # Filter to selected configs
    selected_cfgs = request.form.getlist("configs")
    config_names_in_feedback = sorted({e.config_name for e in entries})
    if not selected_cfgs:
        selected_cfgs = config_names_in_feedback

    # Lazy import so the blueprint loads even without API keys configured
    from src.clients import build_clients
    from src.configs import filter_configs
    from src.runner import run_one

    try:
        configs = filter_configs(selected_cfgs)
    except ValueError as e:
        flash(f"Unknown config: {e}")
        return redirect(url_for("feedback.regression_index"))

    run_id = uuid.uuid4().hex[:8]
    state = {
        "id": run_id,
        "status": "starting",
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "config_names": [c.name for c in configs],
        "total_pairs": len(jobs) * len(configs),
        "completed_pairs": 0,
        "results": [],   # list of InvoiceRegressionResult.__dict__
        "log": [],
        "error": None,
    }
    with _REGRESSION_LOCK:
        _REGRESSION_RUNS[run_id] = state

    def _execute():
        try:
            clients = build_clients(cfg)
            from src.prompts import PromptTemplates
            templates = PromptTemplates.load(cfg["paths"].get("prompts_file", ""))
            for c in configs:
                for job in jobs:
                    fb = fb_index.get((job.invoice_id, c.name))
                    if fb is None:
                        # No feedback for this combo — skip; not regression-relevant.
                        with _REGRESSION_LOCK:
                            state["completed_pairs"] += 1
                        continue
                    try:
                        row = run_one(
                            job, c, clients, cfg, templates=templates
                        )
                    except Exception as e:
                        with _REGRESSION_LOCK:
                            state["log"].append(
                                f"ERROR {c.name} :: {job.invoice_id}: {e}"
                            )
                            state["completed_pairs"] += 1
                        continue
                    pred = row.get("pred") or {}
                    diff = compare(pred, fb)
                    with _REGRESSION_LOCK:
                        state["results"].append({
                            "invoice_id": diff.invoice_id,
                            "config_name": diff.config_name,
                            "previously_correct_still_caught": diff.previously_correct_still_caught,
                            "previously_correct_now_missed": diff.previously_correct_now_missed,
                            "previously_missed_now_caught": diff.previously_missed_now_caught,
                            "previously_missed_still_missed": diff.previously_missed_still_missed,
                            "previously_fp_now_dropped": diff.previously_fp_now_dropped,
                            "previously_fp_still_present": diff.previously_fp_still_present,
                            "new_findings": diff.new_findings,
                            "notes": diff.notes,
                        })
                        state["completed_pairs"] += 1
            with _REGRESSION_LOCK:
                state["status"] = "done"
        except Exception as e:
            with _REGRESSION_LOCK:
                state["status"] = "error"
                state["error"] = f"{type(e).__name__}: {e}"

    threading.Thread(target=_execute, daemon=True).start()
    return redirect(url_for("feedback.regression_view", run_id=run_id))


@bp.route("/regression/run/<run_id>")
def regression_view(run_id: str):
    with _REGRESSION_LOCK:
        state = _REGRESSION_RUNS.get(run_id)
    if not state:
        abort(404)
    # Compute aggregates
    results = state["results"]
    improvements = sum(
        r["previously_missed_now_caught"] + r["previously_fp_now_dropped"]
        for r in results
    )
    regressions = sum(r["previously_correct_now_missed"] for r in results)
    stable = sum(
        r["previously_correct_still_caught"]
        + r["previously_missed_still_missed"]
        + r["previously_fp_still_present"]
        for r in results
    )
    new_findings = sum(r["new_findings"] for r in results)
    return render_template(
        "feedback_regression_view.html",
        state=state,
        improvements=improvements,
        regressions=regressions,
        stable=stable,
        new_findings=new_findings,
    )


@bp.route("/export.xlsx")
def export_xlsx():
    """Export all feedback as Excel for offline review."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    store = _store()
    entries = store.list_all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback Entries"

    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    headers = [
        "feedback_id", "timestamp", "invoice_id", "config_name", "supplier_canonical",
        "supplier_extracted", "invoice_number", "agreement_type", "overall_verdict",
        "agreement_notes", "overall_notes", "line_count", "correct", "false_positive",
        "wrong_amount", "missed_count",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill
        c.font = font
    for e in entries:
        correct = sum(1 for lf in e.line_feedbacks if lf.verdict == "correct")
        fp = sum(1 for lf in e.line_feedbacks if lf.verdict == "false_positive")
        wrong = sum(1 for lf in e.line_feedbacks if lf.verdict == "wrong_amount")
        ws.append([
            e.feedback_id, e.timestamp, e.invoice_id, e.config_name,
            e.supplier_canonical, e.supplier_extracted, e.invoice_number_extracted,
            e.agreement_type, e.overall_verdict, e.agreement_notes, e.overall_notes,
            len(e.line_feedbacks), correct, fp, wrong, len(e.missed_discrepancies),
        ])

    # Detail sheet
    ws2 = wb.create_sheet("Line Verdicts")
    headers2 = [
        "feedback_id", "invoice_id", "config_name", "line_index",
        "verdict", "model_item_number", "model_description",
        "model_quantity", "model_unit_price", "model_agreed_unit_price",
        "model_discrepancy_amount", "correct_unit_price",
        "correct_agreed_unit_price", "correct_discrepancy_amount", "notes",
    ]
    for col, h in enumerate(headers2, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = fill
        c.font = font
    for e in entries:
        for lf in e.line_feedbacks:
            ws2.append([
                e.feedback_id, e.invoice_id, e.config_name, lf.line_index,
                lf.verdict, lf.model_item_number, lf.model_description,
                lf.model_quantity, lf.model_unit_price, lf.model_agreed_unit_price,
                lf.model_discrepancy_amount, lf.correct_unit_price,
                lf.correct_agreed_unit_price, lf.correct_discrepancy_amount, lf.notes,
            ])

    ws3 = wb.create_sheet("Missed Discrepancies")
    headers3 = [
        "feedback_id", "invoice_id", "config_name",
        "item_number", "description", "quantity", "unit_price",
        "correct_agreed_unit_price", "discrepancy_amount", "notes",
    ]
    for col, h in enumerate(headers3, start=1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = fill
        c.font = font
    for e in entries:
        for md in e.missed_discrepancies:
            ws3.append([
                e.feedback_id, e.invoice_id, e.config_name,
                md.item_number, md.description, md.quantity, md.unit_price,
                md.correct_agreed_unit_price, md.discrepancy_amount, md.notes,
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"feedback_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ============================================================================
# AUTOTUNE ROUTES
# ============================================================================

# Background autotune analysis runs (in-memory progress tracker)
_AUTOTUNE_PROGRESS: dict[str, dict] = {}
_AUTOTUNE_LOCK = threading.Lock()


@bp.route("/autotune")
def autotune_index():
    store = _store()
    entries = store.list_all()
    autotune_store = _autotune_store()
    recent = autotune_store.list_recent(limit=10)
    history = _prompt_history()
    versions = history.list_versions()
    versions.sort(key=lambda v: v.timestamp, reverse=True)
    # Cost estimate: rough — meta-prompt ~5k tokens, output ~2k. Sonnet at $3/$15 per M.
    est_input_tokens = 4000 + len(entries) * 200  # base + per-entry summary
    est_cost_usd = (est_input_tokens / 1_000_000) * 3.0 + (2000 / 1_000_000) * 15.0
    return render_template(
        "feedback_autotune.html",
        feedback_count=len(entries),
        recent_runs=recent,
        versions=versions[:10],
        est_input_tokens=est_input_tokens,
        est_cost_usd=est_cost_usd,
    )


@bp.route("/autotune/analyze", methods=["POST"])
def autotune_analyze():
    """Kick off LLM analysis in a background thread; redirect to result page
    that polls progress."""
    store = _store()
    entries = store.list_all()
    if not entries:
        flash("No feedback to analyze. Capture some reviews first.")
        return redirect(url_for("feedback.autotune_index"))

    # Resolve current prompt templates: latest history version if any,
    # otherwise the prompts.json on disk.
    history = _prompt_history()
    latest = history.latest()
    if latest is not None:
        templates = latest.to_templates()
    else:
        from src.prompts import PromptTemplates
        templates = PromptTemplates.load(_prompts_path())

    progress_id = uuid.uuid4().hex[:8]
    with _AUTOTUNE_LOCK:
        _AUTOTUNE_PROGRESS[progress_id] = {
            "status": "running",
            "started_at": datetime.now().isoformat(timespec="seconds"),
            "result_run_id": None,
            "error": None,
        }

    def _execute(app):
        with app.app_context():
            try:
                from src.clients import build_clients
                clients = build_clients(_cfg())
                claude = clients["claude"]
                result = analyze_with_llm(templates, entries, claude_client=claude)
                _autotune_store().save(result)
                with _AUTOTUNE_LOCK:
                    _AUTOTUNE_PROGRESS[progress_id]["status"] = "done"
                    _AUTOTUNE_PROGRESS[progress_id]["result_run_id"] = result.run_id
            except Exception as e:
                with _AUTOTUNE_LOCK:
                    _AUTOTUNE_PROGRESS[progress_id]["status"] = "error"
                    _AUTOTUNE_PROGRESS[progress_id]["error"] = (
                        f"{type(e).__name__}: {e}"
                    )

    threading.Thread(target=_execute, args=(current_app._get_current_object(),), daemon=True).start()
    return redirect(url_for("feedback.autotune_progress", progress_id=progress_id))


@bp.route("/autotune/progress/<progress_id>")
def autotune_progress(progress_id: str):
    with _AUTOTUNE_LOCK:
        state = _AUTOTUNE_PROGRESS.get(progress_id)
    if not state:
        abort(404)
    return render_template(
        "feedback_autotune_progress.html", progress_id=progress_id, state=state
    )


@bp.route("/autotune/result/<run_id>")
def autotune_result(run_id: str):
    result = _autotune_store().get(run_id)
    if not result:
        abort(404)
    # Build the supporting-feedback lookup so we can show inline excerpts
    fb_store = _store()
    fb_by_id = {e.feedback_id: e for e in fb_store.list_all()}
    return render_template(
        "feedback_autotune_result.html",
        result=result,
        fb_by_id=fb_by_id,
    )


@bp.route("/autotune/diagnostics/<run_id>")
def autotune_diagnostics(run_id: str):
    result = _autotune_store().get(run_id)
    if not result:
        abort(404)
    return render_template(
        "feedback_autotune_diagnostics.html", result=result
    )


@bp.route("/autotune/apply/<run_id>/<suggestion_id>", methods=["POST"])
def autotune_apply_one(run_id: str, suggestion_id: str):
    result = _autotune_store().get(run_id)
    if not result:
        abort(404)
    suggestion = next(
        (s for s in result.suggestions if s.suggestion_id == suggestion_id),
        None,
    )
    if suggestion is None:
        flash("Suggestion not found.")
        return redirect(url_for("feedback.autotune_result", run_id=run_id))

    # Optional override of replacement text from the form ("Tweak and apply")
    override = (request.form.get("override_replace") or "").strip()
    if override:
        suggestion.replace = override

    history = _prompt_history()
    latest = history.latest()
    if latest is not None:
        from src.prompts import PromptTemplates
        templates = latest.to_templates()
    else:
        from src.prompts import PromptTemplates
        templates = PromptTemplates.load(_prompts_path())

    try:
        new_templates = apply_suggestion(templates, suggestion)
    except Exception as e:
        flash(f"Could not apply suggestion: {e}")
        return redirect(url_for("feedback.autotune_result", run_id=run_id))

    # Save snapshot
    note = (
        f"autotune: {suggestion.block} — {(suggestion.rationale or '')[:60]}"
    )
    snapshot = PromptVersion.from_templates(
        new_templates,
        source="autotune",
        autotune_run_id=run_id,
        applied_suggestion_ids=[suggestion.suggestion_id],
        note=note,
    )
    history.save_snapshot(snapshot)
    new_templates.save(_prompts_path())
    flash(
        f"Applied suggestion {suggestion.suggestion_id} to '{suggestion.block}'. "
        f"Saved as version {snapshot.version_id}. Consider running a "
        "regression test to verify."
    )
    return redirect(url_for("feedback.autotune_result", run_id=run_id))


@bp.route("/prompt-history")
def prompt_history_page():
    history = _prompt_history()
    versions = history.list_versions()
    versions.sort(key=lambda v: v.timestamp, reverse=True)
    return render_template(
        "feedback_prompt_history.html", versions=versions
    )


@bp.route("/prompt-history/rollback/<version_id>", methods=["POST"])
def prompt_history_rollback(version_id: str):
    history = _prompt_history()
    target = history.get(version_id)
    if not target:
        flash(f"Version {version_id} not found.")
        return redirect(url_for("feedback.prompt_history_page"))
    templates = target.to_templates()
    # Save as a new snapshot tagged as rollback so the chain is auditable
    snapshot = PromptVersion.from_templates(
        templates,
        source="rollback",
        note=f"rollback to {version_id}",
    )
    history.save_snapshot(snapshot)
    templates.save(_prompts_path())
    flash(f"Rolled back to version {version_id}. Saved as new version {snapshot.version_id}.")
    return redirect(url_for("feedback.prompt_history_page"))


@bp.route("/prompt-history/version/<version_id>")
def prompt_history_view(version_id: str):
    history = _prompt_history()
    target = history.get(version_id)
    if not target:
        abort(404)
    return render_template(
        "feedback_prompt_history_view.html", version=target
    )
