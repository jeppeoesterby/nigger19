"""Main evaluation loop."""
from __future__ import annotations

import json
import logging
import threading
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

from openpyxl import load_workbook
from pydantic import ValidationError
from rich.console import Console
from rich.progress import (
    BarColumn,
    Progress,
    SpinnerColumn,
    TextColumn,
    TimeElapsedColumn,
    TimeRemainingColumn,
)

from .clients import ModelCall, parse_model_json
from .configs import ModelConfig
from .excel_writer import build_output_path, json_diff, json_pretty, write_report
from .pricing import cost_usd
from .prompts import (
    PromptTemplates,
    build_hybrid_extraction_prompt,
    build_hybrid_reasoning_prompt,
    build_unified_prompt,
)
from .schema import ExtractedInvoice
from .scorer import InvoiceScores, score_invoice

log = logging.getLogger(__name__)
console = Console()


SUPPORTED_INVOICE_EXTS = {".pdf", ".xlsx"}
SUPPORTED_AGREEMENT_EXTS = {".pdf", ".xlsx"}


@dataclass
class InvoiceJob:
    invoice_id: str
    invoice_path: Path
    agreement_paths: list[Path] = field(default_factory=list)
    ground_truth: Optional[dict] = None  # None when no GT provided


def load_jobs(cfg: dict, limit: Optional[int]) -> list[InvoiceJob]:
    """Discover jobs.

    If ground_truth.json exists, it's the source of truth (which invoices, and
    which agreement each one uses). Otherwise: all files in invoices_dir
    become jobs and every job gets every agreement as context (manual
    verification mode).
    """
    invoices_dir = Path(cfg["paths"]["invoices_dir"])
    agreements_dir = Path(cfg["paths"]["agreements_dir"])
    gt_path = Path(cfg["paths"]["ground_truth"])

    log.info(
        "load_jobs: invoices_dir=%s agreements_dir=%s ground_truth=%s",
        invoices_dir.resolve(),
        agreements_dir.resolve(),
        gt_path.resolve(),
    )

    if not invoices_dir.exists():
        raise FileNotFoundError(
            f"Invoices directory not found: {invoices_dir.resolve()}"
        )

    all_agreements = []
    if agreements_dir.exists():
        for ext in SUPPORTED_AGREEMENT_EXTS:
            all_agreements.extend(sorted(agreements_dir.glob(f"*{ext}")))

    jobs: list[InvoiceJob] = []

    if gt_path.exists():
        ground_truth = json.loads(gt_path.read_text(encoding="utf-8"))
        for invoice_id, gt in ground_truth.items():
            invoice_path = invoices_dir / invoice_id
            if not invoice_path.exists():
                log.warning(
                    "Invoice listed in ground truth but missing on disk: %s",
                    invoice_path,
                )
                continue
            agreement_name = gt.get("agreement_file")
            agreement_paths: list[Path] = []
            if agreement_name:
                ap = agreements_dir / agreement_name
                if ap.exists():
                    agreement_paths.append(ap)
                else:
                    log.warning("Agreement file missing: %s", ap)
            jobs.append(
                InvoiceJob(
                    invoice_id=invoice_id,
                    invoice_path=invoice_path,
                    agreement_paths=agreement_paths,
                    ground_truth=gt,
                )
            )
    else:
        # No ground truth: every file in invoices_dir becomes a job, using all
        # agreements as context. User verifies manually.
        discovered: list[Path] = []
        for ext in SUPPORTED_INVOICE_EXTS:
            discovered.extend(sorted(invoices_dir.glob(f"*{ext}")))
        if not discovered:
            on_disk = [p.name for p in invoices_dir.iterdir() if p.is_file()][:20]
            raise FileNotFoundError(
                f"No invoice files found in {invoices_dir.resolve()} "
                f"(looking for {sorted(SUPPORTED_INVOICE_EXTS)}). "
                f"Files currently in that directory: {on_disk or '(empty)'}. "
                "Upload some invoices first. "
                "Note: on free-tier hosts with ephemeral disks, uploads are "
                "wiped when the service restarts or re-deploys. If that's "
                "likely what happened, just upload again and start the run "
                "immediately after."
            )
        for inv_path in discovered:
            jobs.append(
                InvoiceJob(
                    invoice_id=inv_path.name,
                    invoice_path=inv_path,
                    agreement_paths=list(all_agreements),
                    ground_truth=None,
                )
            )

    if limit is not None:
        jobs = jobs[:limit]
    return jobs


def _xlsx_to_text(path: Path) -> str:
    wb = load_workbook(path, data_only=True, read_only=True)
    out: list[str] = []
    for sheet in wb.worksheets:
        out.append(f"## Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            if all(v is None or v == "" for v in row):
                continue
            out.append("\t".join("" if v is None else str(v) for v in row))
    wb.close()
    return "\n".join(out)


class _FileCache:
    """Per-run cache for file bytes and flattened xlsx text.

    Without this, each invoice PDF is re-read from disk once per model config
    (6x), and each agreement PDF is re-read once per (invoice, config) pair
    (648x for a 108-invoice / 6-config batch). With it, every file reads
    exactly once.
    """

    def __init__(self) -> None:
        self._bytes: dict[str, bytes] = {}
        self._xlsx: dict[str, str] = {}
        self._lock = threading.Lock()

    def read_bytes(self, path: Path) -> bytes:
        key = str(path)
        with self._lock:
            cached = self._bytes.get(key)
            if cached is not None:
                return cached
        data = path.read_bytes()
        with self._lock:
            self._bytes[key] = data
            return data

    def xlsx_text(self, path: Path) -> str:
        key = str(path)
        with self._lock:
            cached = self._xlsx.get(key)
            if cached is not None:
                return cached
        text = _xlsx_to_text(path)
        with self._lock:
            self._xlsx[key] = text
            return text


def _load_invoice_content(
    path: Path, cache: Optional[_FileCache] = None
) -> tuple[Optional[bytes], Optional[str]]:
    """Return (pdf_bytes, xlsx_text). Exactly one is non-None."""
    ext = path.suffix.lower()
    if ext == ".pdf":
        return (cache.read_bytes(path) if cache else path.read_bytes()), None
    if ext == ".xlsx":
        return None, (cache.xlsx_text(path) if cache else _xlsx_to_text(path))
    raise ValueError(f"Unsupported invoice extension: {path}")


def _load_agreement_content(
    paths: list[Path], cache: Optional[_FileCache] = None
) -> tuple[list[bytes], str]:
    """Return (list of PDF bytes, concatenated xlsx text)."""
    pdfs: list[bytes] = []
    text_parts: list[str] = []
    for p in paths:
        ext = p.suffix.lower()
        if ext == ".pdf":
            pdfs.append(cache.read_bytes(p) if cache else p.read_bytes())
        elif ext == ".xlsx":
            text = cache.xlsx_text(p) if cache else _xlsx_to_text(p)
            text_parts.append(f"### {p.name}\n{text}")
    return pdfs, "\n\n".join(text_parts)


def _validate_model_output(raw_text: str) -> tuple[Optional[dict], Optional[str]]:
    if not raw_text or not raw_text.strip():
        return None, "empty response"
    try:
        obj = parse_model_json(raw_text)
    except Exception as e:
        return None, f"json parse: {e}"
    try:
        validated = ExtractedInvoice.model_validate(obj)
    except ValidationError as e:
        return obj, f"schema: {e.error_count()} errors"
    return validated.model_dump(mode="json"), None


def _tok(n) -> int:
    """Defensive int coercion for token counts from SDK responses."""
    try:
        return int(n) if n is not None else 0
    except (TypeError, ValueError):
        return 0


def _lat(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def run_one(
    job: InvoiceJob,
    cfg: ModelConfig,
    clients: dict,
    cfg_yaml: dict,
    templates: Optional[PromptTemplates] = None,
    file_cache: Optional[_FileCache] = None,
) -> dict:
    """Run a single (invoice, config) pair. Always returns a row dict."""
    invoice_pdf, invoice_text = _load_invoice_content(job.invoice_path, file_cache)
    agreement_pdfs, agreement_text = _load_agreement_content(
        job.agreement_paths, file_cache
    )
    has_agreement = bool(agreement_pdfs) or bool(agreement_text)

    total_in = total_out = 0
    total_latency = 0.0
    total_cost = 0.0
    notes: list[str] = []
    pred: Optional[dict] = None
    raw_texts: list[str] = []  # collected model responses for debugging

    def _accumulate(call: ModelCall, model_key: str) -> None:
        nonlocal total_in, total_out, total_latency, total_cost
        total_in += _tok(call.input_tokens)
        total_out += _tok(call.output_tokens)
        total_latency += _lat(call.latency_sec)
        total_cost += cost_usd(
            model_key,
            _tok(call.input_tokens),
            _tok(call.output_tokens),
            cfg_yaml["pricing"],
        )
        # Stash raw text for debug so we can see what the model actually
        # returned, even when parsing failed or output was empty.
        if call.raw_text is not None:
            raw_texts.append(call.raw_text)

    if not cfg.is_hybrid:
        prompt = build_unified_prompt(
            invoice_text=invoice_text,
            agreement_text=agreement_text or None,
            has_invoice_pdf=invoice_pdf is not None,
            agreement_pdf_count=len(agreement_pdfs),
            with_agreement=has_agreement,
            templates=templates,
        )
        # Agreement PDFs are identical across all invoices in this config, so
        # pass them as cached_prefix so Claude can cache them server-side.
        # Invoice PDF is per-call and goes through pdf_documents (not cached).
        per_invoice = [invoice_pdf] if invoice_pdf is not None else []

        client = clients[cfg.extraction.provider]
        call: ModelCall = client.call(
            cfg_yaml["models"][cfg.extraction.model_key],
            prompt,
            pdf_documents=per_invoice,
            cached_prefix_pdfs=agreement_pdfs,
        )
        _accumulate(call, cfg.extraction.model_key)
        if call.error:
            notes.append(f"extract error: {call.error}")
        else:
            pred, err = _validate_model_output(call.raw_text)
            if err:
                notes.append(err)
    else:
        # Stage 1: extraction (invoice only, no agreement)
        extract_prompt = build_hybrid_extraction_prompt(
            invoice_text=invoice_text,
            has_invoice_pdf=invoice_pdf is not None,
            templates=templates,
        )
        extract_attachments = [invoice_pdf] if invoice_pdf is not None else []
        ext_client = clients[cfg.extraction.provider]
        call1 = ext_client.call(
            cfg_yaml["models"][cfg.extraction.model_key],
            extract_prompt,
            extract_attachments,
        )
        _accumulate(call1, cfg.extraction.model_key)
        if call1.error:
            notes.append(f"extract error: {call1.error}")
            extracted: Optional[dict] = None
        else:
            extracted, err = _validate_model_output(call1.raw_text)
            if err:
                notes.append(f"extract {err}")

        if extracted is None:
            pred = None
        elif not has_agreement:
            # Nothing to reason about; return the extracted result as-is.
            pred = extracted
        else:
            # Stage 2: reasoning
            reason_prompt = build_hybrid_reasoning_prompt(
                invoice_json=json.dumps(extracted, ensure_ascii=False, indent=2),
                agreement_text=agreement_text or None,
                agreement_pdf_count=len(agreement_pdfs),
                templates=templates,
            )
            reason_client = clients[cfg.reasoning.provider]
            call2 = reason_client.call(
                cfg_yaml["models"][cfg.reasoning.model_key],
                reason_prompt,
                pdf_documents=[],
                cached_prefix_pdfs=agreement_pdfs,
            )
            _accumulate(call2, cfg.reasoning.model_key)
            if call2.error:
                notes.append(f"reason error: {call2.error}")
                pred = extracted
            else:
                pred2, err2 = _validate_model_output(call2.raw_text)
                if err2:
                    notes.append(f"reason {err2}")
                pred = pred2 if pred2 is not None else extracted

    # Score only if ground truth is available.
    scores: Optional[InvoiceScores] = None
    if job.ground_truth is not None:
        scores = score_invoice(
            job.ground_truth,
            pred or {},
            weights=cfg_yaml["scoring_weights"],
            scoring_cfg=cfg_yaml["scoring"],
        )

    return {
        "invoice_id": job.invoice_id,
        "config_name": cfg.name,
        "field_extraction": scores.field_extraction_pct if scores else None,
        "price_match": scores.price_match_pct if scores else None,
        "credit_note": scores.credit_note_pct if scores else None,
        "rebate": scores.rebate_pct if scores else None,
        "composite": scores.composite if scores else None,
        "latency_sec": total_latency,
        "input_tokens": total_in,
        "output_tokens": total_out,
        "cost_usd": total_cost,
        "notes": "; ".join(notes),
        "per_field": [
            {"field": f.field, "score": f.score}
            for f in (scores.per_field if scores else [])
        ],
        "pred": pred,
        "raw_response": "\n\n--- next call ---\n\n".join(raw_texts) if raw_texts else "",
        "ground_truth": job.ground_truth,
        "failed": pred is None,
        "has_ground_truth": job.ground_truth is not None,
    }


def run(
    cfg_yaml: dict,
    configs: list[ModelConfig],
    clients: dict,
    limit: Optional[int] = None,
    dry_run: bool = False,
    progress_callback: Optional[Callable[[dict], None]] = None,
) -> Optional[Path]:
    jobs = load_jobs(cfg_yaml, limit)
    if not jobs:
        raise RuntimeError(
            "No invoice jobs to run. Upload some invoice PDFs or .xlsx files first."
        )
    total = len(jobs) * len(configs)
    use_web = progress_callback is not None
    has_any_gt = any(j.ground_truth is not None for j in jobs)

    # Snapshot prompt templates at run start so mid-run edits don't mix.
    paths = cfg_yaml.get("paths", {})
    default_prompts = str(
        Path(paths.get("invoices_dir", "data/invoices")).parent / "prompts.json"
    )
    prompts_path = paths.get("prompts_file", default_prompts)
    templates = PromptTemplates.load(prompts_path)

    def emit(event: dict) -> None:
        if progress_callback is not None:
            progress_callback(event)

    emit(
        {
            "type": "plan",
            "jobs": [j.invoice_id for j in jobs],
            "configs": [c.name for c in configs],
            "total": total,
            "scoring_enabled": has_any_gt,
        }
    )
    if not use_web:
        mode = "with ground-truth scoring" if has_any_gt else "extraction-only (no GT)"
        console.print(
            f"[bold]Plan:[/bold] {len(jobs)} invoice(s) x {len(configs)} config(s) = "
            f"{total} total runs ({mode})."
        )
        for c in configs:
            console.print(
                f"  - {c.name:<22} extraction={c.extraction.model_key:<20} "
                f"reasoning={c.reasoning.model_key}"
            )
        console.print("  Invoices:")
        for j in jobs:
            ag = ", ".join(p.name for p in j.agreement_paths) or "none"
            console.print(f"    - {j.invoice_id}  (agreements: {ag})")

    if dry_run:
        emit({"type": "log", "message": "Dry-run: plan validated, no API calls made."})
        emit({"type": "done", "output_path": None})
        if not use_web:
            console.print("[yellow]--dry-run: no API calls made.[/yellow]")
        return None

    per_invoice_rows: list[dict] = []
    raw_rows: list[dict] = []
    emit({"type": "start", "total": total})

    def _record(cfg_name: str, job: InvoiceJob, row: dict) -> None:
        per_invoice_rows.append(row)
        raw_rows.append(
            {
                "invoice_id": job.invoice_id,
                "config_name": cfg_name,
                "ground_truth_json": json_pretty(job.ground_truth) if job.ground_truth else "(no ground truth)",
                "model_output_json": json_pretty(row.get("pred") or {}),
                "raw_response": row.get("raw_response") or "(no response)",
                "diff": json_diff(job.ground_truth or {}, row.get("pred") or {}) if job.ground_truth else "(no ground truth — verify manually)",
            }
        )

    # Per-run PDF/xlsx cache so each file is read from disk exactly once.
    file_cache = _FileCache()

    # Circuit breaker: after CIRCUIT_BREAK_THRESHOLD consecutive failures for a
    # config (typically because the model ID doesn't exist or the API key isn't
    # authorized for that model), stop submitting new work for it. Saves time
    # on deterministic failures. Doesn't cancel in-flight work — but skipped
    # work returns a cheap placeholder row in milliseconds.
    CIRCUIT_BREAK_THRESHOLD = 5
    broken_configs: set[str] = set()
    cfg_counts: dict[str, dict[str, int]] = {}
    cb_lock = threading.Lock()

    def _skipped_row(cfg: ModelConfig, job: InvoiceJob, reason: str) -> dict:
        return {
            "invoice_id": job.invoice_id,
            "config_name": cfg.name,
            "field_extraction": None,
            "price_match": None,
            "credit_note": None,
            "rebate": None,
            "composite": None,
            "latency_sec": 0.0,
            "input_tokens": 0,
            "output_tokens": 0,
            "cost_usd": 0.0,
            "notes": reason,
            "per_field": [],
            "pred": None,
            "raw_response": "",
            "ground_truth": job.ground_truth,
            "failed": True,
            "has_ground_truth": job.ground_truth is not None,
        }

    def _safe_run_one(job: InvoiceJob, cfg: ModelConfig) -> dict:
        """Never raises. On any exception, returns a failure row so the batch continues."""
        # Short-circuit if this config is already tripped.
        with cb_lock:
            if cfg.name in broken_configs:
                return _skipped_row(
                    cfg,
                    job,
                    f"skipped: config hit {CIRCUIT_BREAK_THRESHOLD} consecutive failures",
                )

        try:
            row = run_one(
                job, cfg, clients, cfg_yaml, templates=templates, file_cache=file_cache
            )
        except Exception as e:
            tb = traceback.format_exc()
            log.exception("run_one failed for %s :: %s", cfg.name, job.invoice_id)
            row = {
                "invoice_id": job.invoice_id,
                "config_name": cfg.name,
                "field_extraction": None,
                "price_match": None,
                "credit_note": None,
                "rebate": None,
                "composite": None,
                "latency_sec": 0.0,
                "input_tokens": 0,
                "output_tokens": 0,
                "cost_usd": 0.0,
                "notes": f"internal error: {type(e).__name__}: {e}",
                "per_field": [],
                "pred": None,
                "ground_truth": job.ground_truth,
                "failed": True,
                "has_ground_truth": job.ground_truth is not None,
                "_traceback": tb,
            }

        # Update failure tracker. The breaker should only trip on PERMANENT
        # failures (wrong model ID, auth, bad request). Transient failures
        # like 429 rate limits or 5xx should NOT count — they'll resolve with
        # enough wait time. Reset the consecutive counter on both success AND
        # transient failure so rate-limit churn doesn't kill a healthy config.
        notes_lower = (row.get("notes") or "").lower()
        is_transient_failure = any(
            marker in notes_lower
            for marker in (
                "429",
                "rate_limit",
                "rate limit",
                "timeout",
                "unavailable",
                "temporarily",
                "500 ",
                "502 ",
                "503 ",
                "504 ",
                "connection",
            )
        )
        with cb_lock:
            s = cfg_counts.setdefault(cfg.name, {"consec_fail": 0, "total": 0})
            s["total"] += 1
            if row["failed"] and not is_transient_failure:
                s["consec_fail"] += 1
                if (
                    s["consec_fail"] >= CIRCUIT_BREAK_THRESHOLD
                    and cfg.name not in broken_configs
                ):
                    broken_configs.add(cfg.name)
                    log.warning(
                        "Circuit breaker: disabling %s after %d consecutive permanent failures",
                        cfg.name,
                        s["consec_fail"],
                    )
                    emit(
                        {
                            "type": "log",
                            "message": (
                                f"CIRCUIT BREAKER: {cfg.name} disabled after "
                                f"{s['consec_fail']} consecutive PERMANENT failures "
                                "(not rate limits or timeouts). Remaining invoices "
                                "for this config will be skipped. Likely cause: "
                                "wrong model ID or API key not authorized for this model."
                            ),
                        }
                    )
            else:
                # Success or transient failure -> reset. Don't penalize a
                # working config for rate-limit churn.
                s["consec_fail"] = 0
        return row

    parallelism = max(1, int(cfg_yaml.get("api", {}).get("parallelism", 1) or 1))
    pairs = [(cfg, job) for cfg in configs for job in jobs]
    record_lock = threading.Lock()
    counter = {"n": 0}

    def _do(pair: tuple[ModelConfig, InvoiceJob]) -> tuple[ModelConfig, InvoiceJob, dict]:
        cfg, job = pair
        row = _safe_run_one(job, cfg)
        return cfg, job, row

    if use_web:
        with ThreadPoolExecutor(max_workers=parallelism) as pool:
            futures = [pool.submit(_do, p) for p in pairs]
            for fut in as_completed(futures):
                cfg, job, row = fut.result()
                with record_lock:
                    _record(cfg.name, job, row)
                    counter["n"] += 1
                    completed_n = counter["n"]
                if row.get("_traceback"):
                    emit(
                        {
                            "type": "log",
                            "message": f"TRACEBACK {cfg.name} :: {job.invoice_id}\n{row['_traceback']}",
                        }
                    )
                emit(
                    {
                        "type": "progress",
                        "completed": completed_n,
                        "total": total,
                        "current": f"{cfg.name} :: {job.invoice_id}",
                    }
                )
                emit(
                    {
                        "type": "completed",
                        "completed": completed_n,
                        "total": total,
                        "invoice_id": job.invoice_id,
                        "config_name": cfg.name,
                        "composite": row["composite"],
                        "notes": row["notes"],
                        "has_ground_truth": row["has_ground_truth"],
                    }
                )
    else:
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("{task.completed}/{task.total}"),
            TimeElapsedColumn(),
            TimeRemainingColumn(),
            console=console,
        ) as progress:
            task = progress.add_task("Evaluating", total=total)
            with ThreadPoolExecutor(max_workers=parallelism) as pool:
                futures = [pool.submit(_do, p) for p in pairs]
                for fut in as_completed(futures):
                    cfg, job, row = fut.result()
                    with record_lock:
                        _record(cfg.name, job, row)
                    progress.update(task, description=f"{cfg.name} :: {job.invoice_id}")
                    progress.advance(task)

    summary_rows = _summarize(per_invoice_rows, configs, has_any_gt)
    out_path = build_output_path(cfg_yaml["paths"]["results_dir"])
    write_report(
        out_path,
        summary_rows,
        per_invoice_rows,
        raw_rows,
        scoring_enabled=has_any_gt,
    )
    emit(
        {
            "type": "done",
            "output_path": str(out_path),
            "summary": summary_rows,
            "scoring_enabled": has_any_gt,
        }
    )
    if not use_web:
        console.print(f"[green]Wrote[/green] {out_path}")
    return out_path


def _summarize(
    per_invoice_rows: list[dict],
    configs: list[ModelConfig],
    has_any_gt: bool,
) -> list[dict]:
    summary: list[dict] = []
    for cfg in configs:
        rows = [r for r in per_invoice_rows if r["config_name"] == cfg.name]
        if not rows:
            continue
        n = len(rows)
        failures = sum(1 for r in rows if r["failed"])
        scored = [r for r in rows if not r["failed"] and r["has_ground_truth"]]
        ns = len(scored)

        def _avg(field: str) -> Optional[float]:
            if not ns:
                return None
            return sum(r[field] for r in scored) / ns

        total_cost = sum(r["cost_usd"] for r in rows)
        total_in = sum(r["input_tokens"] for r in rows)
        total_out = sum(r["output_tokens"] for r in rows)
        avg_lat = sum(r["latency_sec"] for r in rows) / n if n else 0.0

        summary.append(
            {
                "config_name": cfg.name,
                "docs_tested": n,
                "field_extraction_accuracy": _avg("field_extraction"),
                "price_match_accuracy": _avg("price_match"),
                "credit_note_accuracy": _avg("credit_note"),
                "rebate_accuracy": _avg("rebate"),
                "composite_score": _avg("composite"),
                "avg_latency_sec": avg_lat,
                "total_input_tokens": total_in,
                "total_output_tokens": total_out,
                "total_cost_usd": total_cost,
                "cost_per_doc_usd": total_cost / n if n else 0.0,
                "failures": failures,
            }
        )
    return summary
