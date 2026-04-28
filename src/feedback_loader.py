"""Load past run results from the Excel report for feedback review.

Rather than modify runner.py to write a JSON sidecar, we re-read the
existing eval_*.xlsx file. The Raw-Outputs / Model-Outputs sheet
contains the full model_output_json per (invoice, config) — enough to
pre-populate the feedback review form.

Strictly additive: nothing in runner or excel_writer is touched.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


@dataclass
class RunPair:
    """One (invoice, config) cell from a past run."""

    invoice_id: str
    config_name: str
    model_pred: dict        # parsed JSON of what model returned
    raw_response: str       # raw text (debugging only)
    notes: str              # error/warning notes from runner
    has_ground_truth: bool


def list_runs(results_dir: Path) -> list[dict]:
    """List eval_*.xlsx files newest first."""
    if not results_dir.exists():
        return []
    rows = []
    for p in sorted(results_dir.glob("eval_*.xlsx"), reverse=True):
        st = p.stat()
        rows.append(
            {
                "filename": p.name,
                "size_kb": max(1, st.st_size // 1024),
                "mtime": p.stat().st_mtime,
            }
        )
    return rows


def load_run(xlsx_path: Path) -> list[RunPair]:
    """Read a past evaluation Excel and return one RunPair per (invoice, config)."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    pairs = _load_from_raw_sheet(wb)
    if not pairs:
        # If for some reason the raw sheet is missing, fall back to Per-Invoice.
        pairs = _load_from_per_invoice(wb)
    wb.close()
    return pairs


def _find_sheet(wb, *candidates: str):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    return None


def _load_from_raw_sheet(wb) -> list[RunPair]:
    ws = _find_sheet(wb, "Raw-Outputs", "Model-Outputs (manual review)")
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h or "").strip().lower() for h in rows[0]]

    def col(*names: str) -> Optional[int]:
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    inv_idx = col("invoice_id")
    cfg_idx = col("config_name")
    json_idx = col("model_output_json")
    raw_idx = col("raw_response_text", "raw_response")
    if inv_idx is None or cfg_idx is None or json_idx is None:
        return []

    out: list[RunPair] = []
    for row in rows[1:]:
        invoice_id = str(row[inv_idx] or "").strip()
        config_name = str(row[cfg_idx] or "").strip()
        if not invoice_id or not config_name:
            continue
        pred = _parse_json_safe(row[json_idx])
        raw = str(row[raw_idx] or "") if raw_idx is not None else ""
        out.append(
            RunPair(
                invoice_id=invoice_id,
                config_name=config_name,
                model_pred=pred,
                raw_response=raw,
                notes="",
                has_ground_truth=False,
            )
        )
    return out


def _load_from_per_invoice(wb) -> list[RunPair]:
    ws = _find_sheet(wb, "Per-Invoice")
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h or "").strip().lower() for h in rows[0]]

    def col(*names: str) -> Optional[int]:
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    inv_idx = col("invoice_id")
    cfg_idx = col("config_name")
    notes_idx = col("notes")
    if inv_idx is None or cfg_idx is None:
        return []

    out: list[RunPair] = []
    for row in rows[1:]:
        invoice_id = str(row[inv_idx] or "").strip()
        config_name = str(row[cfg_idx] or "").strip()
        if not invoice_id or not config_name:
            continue
        out.append(
            RunPair(
                invoice_id=invoice_id,
                config_name=config_name,
                model_pred={},
                raw_response="",
                notes=str(row[notes_idx] or "") if notes_idx is not None else "",
                has_ground_truth=False,
            )
        )
    return out


def _parse_json_safe(s) -> dict:
    if s is None:
        return {}
    text = str(s).strip()
    if not text:
        return {}
    try:
        obj = json.loads(text)
    except json.JSONDecodeError:
        # The Excel cell may have been clipped / re-wrapped. Try to extract
        # the first JSON object.
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if not m:
            return {}
        try:
            obj = json.loads(m.group(0))
        except json.JSONDecodeError:
            return {}
    return obj if isinstance(obj, dict) else {}
