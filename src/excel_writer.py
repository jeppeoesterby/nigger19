"""Write the 4-sheet evaluation Excel report."""
from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, headers: list[str], max_width: int = 60, sample_rows: int = 50) -> None:
    """Set column widths. For large sheets (>sample_rows), sample the first N
    rows rather than iterating all — avoids O(n) per column on sheets with
    thousands of rows. The cost of imprecise column width on deeper rows is
    negligible vs. the perf hit of full iteration."""
    max_row_to_check = min(ws.max_row, sample_rows + 1)
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        longest = len(str(header))
        if max_row_to_check >= 2:
            for row in ws.iter_rows(
                min_row=2, max_row=max_row_to_check, min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    v = cell.value
                    if v is None:
                        continue
                    s = str(v)
                    if "\n" in s:
                        s = s.split("\n")[0]
                    longest = max(longest, len(s))
        ws.column_dimensions[col_letter].width = min(max_width, longest + 2)


def _pct(v) -> str:
    """Format fraction as percent string, or em-dash if None."""
    if v is None:
        return "—"
    return f"{round(v * 100, 2)}"


def build_output_path(results_dir: str) -> Path:
    ts = datetime.now().strftime("%Y-%m-%d_%H%M")
    p = Path(results_dir) / f"eval_{ts}.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


def write_report(
    output_path: Path,
    summary_rows: list[dict],
    per_invoice_rows: list[dict],
    raw_rows: list[dict],
    *,
    scoring_enabled: bool = True,
) -> None:
    wb = Workbook()

    # Sheet 1: Fejloversigt — flat audit table matching the manual report format
    # (one row per discrepant line item). config_name is the first column so
    # users can compare models via Excel auto-filter or sort.
    ws = wb.active
    ws.title = "Fejloversigt"
    _write_fejloversigt(ws, per_invoice_rows)

    # Sheet 2: Benchmark — pivot view (one row per unique error, one column per
    # model). Lets the user see at a glance which models caught which errors —
    # consensus errors at the top, controversial ones below.
    ws = wb.create_sheet("Benchmark")
    _write_benchmark(ws, per_invoice_rows)

    # Sheet 3: Diagnostik — per-pair "WHY" view. One row per (invoice, config)
    # showing extraction status, line counts, error notes, raw-response excerpt
    # so a user reviewing missed errors can see WHY (e.g. extraction failed,
    # no agreement match found, JSON parse failed, empty response).
    ws = wb.create_sheet("Diagnostik")
    _write_diagnostics(ws, per_invoice_rows)

    # Sheet 4: Alle linjer — every extracted line across all pairs, with match
    # status. Lets the user filter "show me all lines for INV-001 across all
    # models" or "show me lines without agreement match" to debug deeper.
    ws = wb.create_sheet("Alle linjer")
    _write_all_lines(ws, per_invoice_rows)

    # Sheet 5: Findings (human-readable per-invoice analysis)
    # Keeps the per-invoice context (totals, rebate, credit-note status) that
    # the flat Fejloversigt deliberately omits.
    ws = wb.create_sheet("Findings")
    headers = [
        "invoice_id",
        "config_name",
        "supplier",
        "invoice_number",
        "invoice_date",
        "document_type",
        "currency",
        "total",
        "line_count",
        "discrepancy_count",
        "discrepancies_detail",
        "rebate_status",
        "credit_note_info",
        "status",
    ]
    _write_header(ws, headers)
    for r in per_invoice_rows:
        pred = r.get("pred") or {}
        notes = (r.get("notes") or "").strip()
        line_items = pred.get("line_items") or []
        discrepancies = [li for li in line_items if li.get("has_discrepancy")]
        discrepancy_text = _format_discrepancies(discrepancies) if discrepancies else (
            "(ingen uoverensstemmelser fundet)" if line_items else "(ingen line items)"
        )
        ws.append(
            [
                r.get("invoice_id"),
                r.get("config_name"),
                pred.get("supplier_name") or "—",
                pred.get("invoice_number") or "—",
                pred.get("invoice_date") or "—",
                pred.get("document_type") or "—",
                pred.get("currency") or "—",
                pred.get("total") if pred.get("total") is not None else "—",
                len(line_items),
                len(discrepancies),
                discrepancy_text,
                _format_rebate(pred),
                _format_credit_note(pred),
                "OK" if not notes else f"FEJL: {notes}",
            ]
        )
    # Wide cols for the multi-line text columns + wrap-text
    for col_idx in (11, 12, 13, 14):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 60
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    # Sensible widths for the short cols
    short_widths = {1: 24, 2: 18, 3: 22, 4: 16, 5: 12, 6: 14, 7: 8, 8: 10, 9: 6, 10: 6}
    for idx, w in short_widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Sheet 2: Summary
    ws = wb.create_sheet("Summary")
    if scoring_enabled:
        headers = [
            "config_name",
            "docs_tested",
            "field_extraction_accuracy_%",
            "price_match_accuracy_%",
            "credit_note_accuracy_%",
            "rebate_accuracy_%",
            "composite_score_%",
            "avg_latency_sec",
            "total_input_tokens",
            "total_output_tokens",
            "cache_create_tokens",
            "cache_read_tokens",
            "total_cost_usd",
            "cost_per_doc_usd",
            "failures",
        ]
    else:
        headers = [
            "config_name",
            "docs_tested",
            "avg_latency_sec",
            "total_input_tokens",
            "total_output_tokens",
            "cache_create_tokens",
            "cache_read_tokens",
            "total_cost_usd",
            "cost_per_doc_usd",
            "failures",
        ]
    _write_header(ws, headers)
    for r in summary_rows:
        cc = r.get("total_cache_creation_tokens", 0)
        cr = r.get("total_cache_read_tokens", 0)
        if scoring_enabled:
            ws.append(
                [
                    r["config_name"],
                    r["docs_tested"],
                    _pct(r["field_extraction_accuracy"]),
                    _pct(r["price_match_accuracy"]),
                    _pct(r["credit_note_accuracy"]),
                    _pct(r["rebate_accuracy"]),
                    _pct(r["composite_score"]),
                    round(r["avg_latency_sec"], 2),
                    r["total_input_tokens"],
                    r["total_output_tokens"],
                    cc,
                    cr,
                    round(r["total_cost_usd"], 4),
                    round(r["cost_per_doc_usd"], 4),
                    r["failures"],
                ]
            )
        else:
            ws.append(
                [
                    r["config_name"],
                    r["docs_tested"],
                    round(r["avg_latency_sec"], 2),
                    r["total_input_tokens"],
                    r["total_output_tokens"],
                    cc,
                    cr,
                    round(r["total_cost_usd"], 4),
                    round(r["cost_per_doc_usd"], 4),
                    r["failures"],
                ]
            )
    _autosize(ws, headers)

    # Sheet 2: Per-Invoice
    ws = wb.create_sheet("Per-Invoice")
    if scoring_enabled:
        headers = [
            "invoice_id",
            "config_name",
            "field_extraction_%",
            "price_match_%",
            "credit_note_%",
            "rebate_%",
            "composite_%",
            "latency_sec",
            "input_tokens",
            "output_tokens",
            "cost_usd",
            "notes",
        ]
    else:
        headers = [
            "invoice_id",
            "config_name",
            "latency_sec",
            "input_tokens",
            "output_tokens",
            "cost_usd",
            "notes",
        ]
    _write_header(ws, headers)
    for r in per_invoice_rows:
        if scoring_enabled:
            ws.append(
                [
                    r["invoice_id"],
                    r["config_name"],
                    _pct(r["field_extraction"]),
                    _pct(r["price_match"]),
                    _pct(r["credit_note"]),
                    _pct(r["rebate"]),
                    _pct(r["composite"]),
                    round(r["latency_sec"], 2),
                    r["input_tokens"],
                    r["output_tokens"],
                    round(r["cost_usd"], 4),
                    r.get("notes", ""),
                ]
            )
        else:
            ws.append(
                [
                    r["invoice_id"],
                    r["config_name"],
                    round(r["latency_sec"], 2),
                    r["input_tokens"],
                    r["output_tokens"],
                    round(r["cost_usd"], 4),
                    r.get("notes", ""),
                ]
            )
    _autosize(ws, headers)

    # Sheet 3: Per-Field (aggregated from per_invoice_rows[...]["per_field"])
    # Only meaningful when scoring is enabled.
    if scoring_enabled:
        ws = wb.create_sheet("Per-Field")
        headers = [
            "config_name",
            "field",
            "total_attempts",
            "exact_match",
            "partial_match",
            "wrong",
            "accuracy_pct",
        ]
        _write_header(ws, headers)
        agg: dict[tuple[str, str], dict] = defaultdict(
            lambda: {"attempts": 0, "exact": 0, "partial": 0, "wrong": 0, "sum": 0.0}
        )
        for row in per_invoice_rows:
            cfg = row["config_name"]
            for f in row.get("per_field", []):
                k = (cfg, f["field"])
                agg[k]["attempts"] += 1
                s = f["score"]
                agg[k]["sum"] += s
                if s >= 1.0:
                    agg[k]["exact"] += 1
                elif s > 0:
                    agg[k]["partial"] += 1
                else:
                    agg[k]["wrong"] += 1
        for (cfg, fname), d in sorted(agg.items()):
            acc = d["sum"] / d["attempts"] if d["attempts"] else 0.0
            ws.append(
                [cfg, fname, d["attempts"], d["exact"], d["partial"], d["wrong"], _pct(acc)]
            )
        _autosize(ws, headers)

    # Sheet 4: Raw-Outputs (this is the main sheet when no GT)
    raw_sheet_name = (
        "Raw-Outputs" if scoring_enabled else "Model-Outputs (manual review)"
    )
    ws = wb.create_sheet(raw_sheet_name)
    if scoring_enabled:
        headers = [
            "invoice_id",
            "config_name",
            "ground_truth_json",
            "model_output_json",
            "raw_response_text",
            "diff",
        ]
    else:
        headers = [
            "invoice_id",
            "config_name",
            "model_output_json",
            "raw_response_text",
        ]
    _write_header(ws, headers)
    for r in raw_rows:
        if scoring_enabled:
            ws.append(
                [
                    r["invoice_id"],
                    r["config_name"],
                    r["ground_truth_json"],
                    r["model_output_json"],
                    _clip_for_excel(r.get("raw_response") or "(no response)"),
                    r["diff"],
                ]
            )
        else:
            ws.append(
                [
                    r["invoice_id"],
                    r["config_name"],
                    r["model_output_json"],
                    _clip_for_excel(r.get("raw_response") or "(no response)"),
                ]
            )
    # JSON + raw-text columns: wrap text
    text_cols = [3, 4, 5, 6] if scoring_enabled else [3, 4]
    for col_idx in text_cols:
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 70
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_idx in (1, 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

    wb.save(output_path)


def _clip_for_excel(s: str, max_len: int = 32000) -> str:
    """Excel cap on a single cell is ~32767 chars. Clip to be safe."""
    if s is None:
        return ""
    if len(s) <= max_len:
        return s
    return s[:max_len] + f"\n... [clipped {len(s) - max_len} chars]"


def json_pretty(obj: Any) -> str:
    return json.dumps(obj, indent=2, ensure_ascii=False, default=str)


SUBTOTAL_FILL = PatternFill(start_color="E5F0EA", end_color="E5F0EA", fill_type="solid")
SUBTOTAL_FONT = Font(bold=True, color="0E7C2E")
GRAND_TOTAL_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
GRAND_TOTAL_FONT = Font(bold=True, color="FFFFFF", size=12)


def _write_fejloversigt(ws, per_invoice_rows: list[dict]) -> None:
    """Audit-style sheet: one row per discrepant line item, grouped by config.

    Layout matches the manual "Fejloversigt" report — Faktura Nummer,
    Varenummer, Faktisk købspris, Rabatpris, Antal, Faktisk total pris,
    Total rabatpris, Sum overbetalt — with config_name as the leading column
    so the user can compare models side-by-side via Excel's auto-filter.

    A subtotal row sits at the end of each model block and a grand-total
    comparison block at the very top makes "which model caught the most
    overcharges" obvious.
    """
    # Build flat list of discrepant lines, grouped by config preserving the
    # order configs appear in per_invoice_rows.
    by_config: dict[str, list[dict]] = {}
    config_order: list[str] = []
    for r in per_invoice_rows:
        cfg = r.get("config_name") or "(unknown)"
        if cfg not in by_config:
            by_config[cfg] = []
            config_order.append(cfg)
        pred = r.get("pred") or {}
        invoice_no = pred.get("invoice_number") or r.get("invoice_id") or ""
        for li in pred.get("line_items") or []:
            if not li.get("has_discrepancy"):
                continue
            qty = _coerce_float(li.get("quantity"))
            unit = _coerce_float(li.get("unit_price"))
            agreed = _coerce_float(li.get("agreed_unit_price"))
            line_total = _coerce_float(li.get("line_total"))
            faktisk_total = line_total if line_total is not None else (
                round(unit * qty, 2) if (unit is not None and qty is not None) else None
            )
            agreed_total = (
                round(agreed * qty, 2) if (agreed is not None and qty is not None) else None
            )
            sum_over = _coerce_float(li.get("discrepancy_amount"))
            if sum_over is None and faktisk_total is not None and agreed_total is not None:
                sum_over = round(faktisk_total - agreed_total, 2)
            by_config[cfg].append(
                {
                    "faktura_nummer": str(invoice_no),
                    "varenummer": (li.get("item_number") or "").strip() if li.get("item_number") else "",
                    "beskrivelse": (li.get("description") or "").strip(),
                    "faktisk_købspris": unit,
                    "rabatpris": agreed,
                    "antal": qty,
                    "faktisk_total": faktisk_total,
                    "total_rabatpris": agreed_total,
                    "sum_overbetalt": sum_over,
                }
            )

    # --- Top: Model-sammenligning block ---
    title = ws.cell(row=1, column=1, value="Fejloversigt — model-sammenligning")
    title.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    cmp_headers = [
        "Model",
        "# uoverensstemmelser",
        "Total sum overbetalt (DKK)",
    ]
    for col, h in enumerate(cmp_headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    row = 4
    for cfg in config_order:
        items = by_config[cfg]
        total_over = sum((x["sum_overbetalt"] or 0) for x in items)
        ws.cell(row=row, column=1, value=cfg)
        ws.cell(row=row, column=2, value=len(items))
        ws.cell(row=row, column=3, value=round(total_over, 2))
        row += 1

    # --- Detail block ---
    detail_start_row = row + 2
    title2 = ws.cell(row=detail_start_row, column=1, value="Detaljer pr. uoverensstemmelse")
    title2.font = Font(bold=True, size=12)
    ws.merge_cells(
        start_row=detail_start_row,
        start_column=1,
        end_row=detail_start_row,
        end_column=10,
    )

    headers = [
        "Model",
        "Faktura Nummer",
        "Varenummer",
        "Beskrivelse",
        "Faktisk købspris",
        "Rabatpris",
        "Antal",
        "Faktisk total pris",
        "Total rabatpris",
        "Sum overbetalt",
    ]
    header_row_idx = detail_start_row + 2
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row_idx, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")

    row = header_row_idx + 1
    grand_total = 0.0
    for cfg in config_order:
        items = by_config[cfg]
        if not items:
            ws.cell(row=row, column=1, value=cfg)
            ws.cell(
                row=row, column=2,
                value="(ingen uoverensstemmelser fundet for denne model)",
            )
            row += 1
            continue
        for it in items:
            ws.cell(row=row, column=1, value=cfg)
            ws.cell(row=row, column=2, value=it["faktura_nummer"])
            ws.cell(row=row, column=3, value=it["varenummer"])
            ws.cell(row=row, column=4, value=it["beskrivelse"][:120])
            ws.cell(row=row, column=5, value=it["faktisk_købspris"])
            ws.cell(row=row, column=6, value=it["rabatpris"])
            ws.cell(row=row, column=7, value=it["antal"])
            ws.cell(row=row, column=8, value=it["faktisk_total"])
            ws.cell(row=row, column=9, value=it["total_rabatpris"])
            ws.cell(row=row, column=10, value=it["sum_overbetalt"])
            for col in (5, 6, 8, 9, 10):
                ws.cell(row=row, column=col).number_format = '#,##0.00 "kr."'
            row += 1
        # Subtotal row for this config
        subtotal = round(sum((x["sum_overbetalt"] or 0) for x in items), 2)
        grand_total += subtotal
        sub_label = ws.cell(
            row=row, column=1, value=f"Samlet for {cfg} (manglende rabat):"
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        sub_label.fill = SUBTOTAL_FILL
        sub_label.font = SUBTOTAL_FONT
        sub_label.alignment = Alignment(horizontal="right", vertical="center")
        sub_val = ws.cell(row=row, column=10, value=subtotal)
        sub_val.fill = SUBTOTAL_FILL
        sub_val.font = SUBTOTAL_FONT
        sub_val.number_format = '#,##0.00 "kr."'
        row += 2  # blank row before next config

    # Auto-filter on the detail header so users can filter by Model
    ws.auto_filter.ref = (
        f"A{header_row_idx}:J{max(header_row_idx, row - 1)}"
    )
    # Freeze the detail header row
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1).coordinate

    # Column widths
    widths = {1: 22, 2: 16, 3: 16, 4: 38, 5: 16, 6: 14, 7: 8, 8: 18, 9: 16, 10: 16}
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def _classify_status(notes: str, pred: dict | None) -> str:
    """Classify the runner's `notes` string into a coarse status code.

    Used in the Diagnostik sheet so the user can filter / sort by failure
    mode. Order matters — earlier patterns win when notes contain multiple.
    """
    n = (notes or "").lower()
    if "circuit breaker" in n or "skipped: config" in n:
        return "CIRCUIT_BROKEN"
    if "internal error" in n:
        return "INTERNAL_ERROR"
    if "extract error" in n:
        if "429" in n or "rate" in n:
            return "EXTRACT_RATE_LIMITED"
        if "timeout" in n:
            return "EXTRACT_TIMEOUT"
        return "EXTRACT_FAILED"
    if "reason error" in n:
        return "REASONING_FAILED"
    if "json parse" in n or "json_parse" in n:
        return "JSON_PARSE_FAILED"
    if "schema" in n and "errors" in n:
        return "SCHEMA_VALIDATION_FAILED"
    if "empty response" in n or "[empty response debug]" in n:
        return "EMPTY_RESPONSE"
    if pred is None or not pred.get("line_items"):
        # No structured output produced
        if not n:
            return "NO_OUTPUT"
        return "WARNING"
    if not n:
        return "OK"
    return "WARNING"


def _write_diagnostics(ws, per_invoice_rows: list[dict]) -> None:
    """Per (invoice, config) row explaining WHY the model produced what it did.

    Surfaces extraction status, line counts (matched / unmatched / discrepant),
    token usage, latency, cost, error notes, and a raw-response excerpt so a
    user reviewing a missed error can immediately see whether the cause is:
      - extraction failed / rate-limited / timed out
      - JSON didn't parse
      - schema validation failed
      - empty response from the model
      - lines extracted but didn't match the agreement
      - circuit breaker tripped
      - or just OK (the model genuinely didn't see a discrepancy)
    """
    headers = [
        "invoice_id",
        "config_name",
        "status",                         # OK / EXTRACT_FAILED / etc.
        "supplier",
        "invoice_number",
        "total_lines",
        "discrepancies",
        "matched_no_discrepancy",
        "unmatched_to_agreement",
        "not_evaluated",
        "input_tokens",
        "output_tokens",
        "latency_sec",
        "cost_usd",
        "notes",
        "raw_response_excerpt",
    ]
    _write_header(ws, headers)
    for r in per_invoice_rows:
        pred = r.get("pred") or {}
        notes = r.get("notes") or ""
        status = _classify_status(notes, pred)
        line_items = pred.get("line_items") or []
        total = len(line_items)
        disc = sum(1 for li in line_items if li.get("has_discrepancy") is True)
        matched_no_disc = sum(
            1
            for li in line_items
            if li.get("has_discrepancy") is False
            and li.get("agreed_unit_price") is not None
        )
        unmatched = sum(
            1
            for li in line_items
            if li.get("agreed_unit_price") is None and li.get("has_discrepancy") is not None
        )
        not_evaluated = sum(
            1 for li in line_items if li.get("has_discrepancy") is None
        )
        raw = r.get("raw_response") or ""
        ws.append(
            [
                r.get("invoice_id"),
                r.get("config_name"),
                status,
                pred.get("supplier_name") or "—",
                pred.get("invoice_number") or "—",
                total,
                disc,
                matched_no_disc,
                unmatched,
                not_evaluated,
                int(r.get("input_tokens") or 0),
                int(r.get("output_tokens") or 0),
                round(float(r.get("latency_sec") or 0), 2),
                round(float(r.get("cost_usd") or 0), 4),
                notes,
                _clip_for_excel(raw, max_len=2000),
            ]
        )

    # Auto-filter on header so users can filter by status
    last_col = get_column_letter(len(headers))
    last_row = max(2, ws.max_row)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = "A2"

    widths = {
        1: 24, 2: 22, 3: 24, 4: 22, 5: 16,
        6: 8, 7: 8, 8: 10, 9: 10, 10: 8,
        11: 10, 12: 10, 13: 10, 14: 10,
        15: 50, 16: 60,
    }
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w
    # Wrap text on notes + raw_response_excerpt
    for col_idx in (15, 16):
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_all_lines(ws, per_invoice_rows: list[dict]) -> None:
    """One row per extracted line item (across all invoices and configs).

    Lets the user filter to e.g. "show me all unmatched lines from
    gemini-pure" to see which products the model couldn't find in the
    agreement — often the reason it missed a discrepancy.
    """
    headers = [
        "invoice_id",
        "config_name",
        "line_idx",
        "item_number",
        "description",
        "quantity",
        "unit_price",
        "agreed_unit_price",
        "line_total",
        "has_discrepancy",
        "discrepancy_amount",
        "match_status",
    ]
    _write_header(ws, headers)
    for r in per_invoice_rows:
        pred = r.get("pred") or {}
        for i, li in enumerate(pred.get("line_items") or []):
            agreed = li.get("agreed_unit_price")
            has_disc = li.get("has_discrepancy")
            if has_disc is True:
                status = "discrepancy"
            elif has_disc is False and agreed is not None:
                status = "matched_no_discrepancy"
            elif agreed is None and has_disc is not None:
                status = "unmatched"
            else:
                status = "not_evaluated"
            ws.append(
                [
                    r.get("invoice_id"),
                    r.get("config_name"),
                    i,
                    (li.get("item_number") or ""),
                    (li.get("description") or "")[:120],
                    li.get("quantity"),
                    li.get("unit_price"),
                    li.get("agreed_unit_price"),
                    li.get("line_total"),
                    has_disc if has_disc is not None else "",
                    li.get("discrepancy_amount"),
                    status,
                ]
            )

    last_col = get_column_letter(len(headers))
    last_row = max(2, ws.max_row)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    ws.freeze_panes = "A2"

    for col in (7, 8, 9, 11):
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.number_format = '#,##0.00 "kr."'

    widths = {1: 22, 2: 22, 3: 6, 4: 16, 5: 40, 6: 8, 7: 12, 8: 14, 9: 12, 10: 12, 11: 14, 12: 22}
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def _write_benchmark(ws, per_invoice_rows: list[dict]) -> None:
    """Pivot view: one row per unique error, one column per model.

    Dedup key: (invoice_number, item_number) when item_number is present;
    otherwise (invoice_number, lowercased description). Imperfect matching
    across models is itself informative — if two rows look "the same" but
    weren't merged, it means models extracted different keys, which the
    user should know about.

    Sort order: errors caught by all models first (likely real), then by
    max sum_overbetalt across models. Controversial errors (caught by
    only some models) sink to the middle/bottom for inspection.
    """
    # Collect models in first-seen order
    model_order: list[str] = []
    seen: set[str] = set()
    for r in per_invoice_rows:
        cfg = r.get("config_name") or "(unknown)"
        if cfg not in seen:
            seen.add(cfg)
            model_order.append(cfg)

    # Build pivot map keyed by error signature
    by_key: dict[tuple, dict] = {}
    for r in per_invoice_rows:
        cfg = r.get("config_name") or "(unknown)"
        pred = r.get("pred") or {}
        invoice_no = pred.get("invoice_number") or r.get("invoice_id") or ""
        for li in pred.get("line_items") or []:
            if not li.get("has_discrepancy"):
                continue
            varenr = (li.get("item_number") or "").strip() if li.get("item_number") else ""
            desc = (li.get("description") or "").strip()
            sig = varenr or desc.lower()
            if not sig:
                # Skip malformed rows with no description and no item_number;
                # they can't be deduped meaningfully.
                continue
            key = (str(invoice_no), sig)
            qty = _coerce_float(li.get("quantity"))
            unit = _coerce_float(li.get("unit_price"))
            agreed = _coerce_float(li.get("agreed_unit_price"))
            sum_over = _coerce_float(li.get("discrepancy_amount"))
            entry = by_key.setdefault(
                key,
                {
                    "faktura": invoice_no,
                    "varenummer": varenr,
                    "beskrivelse": desc,
                    "antal": qty,
                    "unit_price": unit,
                    "agreed_unit_price": agreed,
                    "by_model": {},
                },
            )
            # First non-null value wins for the consensus columns. Different
            # models reporting different prices on the same key is rare but
            # possible; the by_model column shows what each one said.
            if entry["antal"] is None and qty is not None:
                entry["antal"] = qty
            if entry["unit_price"] is None and unit is not None:
                entry["unit_price"] = unit
            if entry["agreed_unit_price"] is None and agreed is not None:
                entry["agreed_unit_price"] = agreed
            if sum_over is not None:
                entry["by_model"][cfg] = sum_over

    n_models = len(model_order)
    total_unique_errors = len(by_key)

    # --- Top: per-model summary ---
    title = ws.cell(row=1, column=1, value="Benchmark — pr. unik fejl, pr. model")
    title.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(8, 4 + n_models))

    summary_headers = [
        "Model",
        "Fejl fundet",
        "Coverage",
        "Total sum overbetalt (DKK)",
    ]
    for col, h in enumerate(summary_headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    row = 4
    for m in model_order:
        caught = sum(1 for e in by_key.values() if m in e["by_model"])
        total_sum = round(sum(e["by_model"].get(m, 0) for e in by_key.values()), 2)
        coverage = f"{caught}/{total_unique_errors}" if total_unique_errors else "—"
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=caught)
        ws.cell(row=row, column=3, value=coverage)
        c_total = ws.cell(row=row, column=4, value=total_sum)
        c_total.number_format = '#,##0.00 "kr."'
        row += 1

    # --- Detail pivot ---
    detail_start = row + 2
    title2 = ws.cell(
        row=detail_start, column=1, value="Pivot — én række pr. unik fejl"
    )
    title2.font = Font(bold=True, size=12)
    ws.merge_cells(
        start_row=detail_start,
        start_column=1,
        end_row=detail_start,
        end_column=max(8, 7 + n_models),
    )

    fixed_headers = [
        "Faktura Nummer",
        "Varenummer",
        "Beskrivelse",
        "Antal",
        "Faktisk købspris",
        "Rabatpris (aftalt)",
    ]
    headers = fixed_headers + model_order + ["Found by"]
    header_row = detail_start + 2
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    # Sort: most-caught first (consensus = real errors), then by largest
    # max-amount across models (so big-money items surface).
    def _sort_key(e: dict):
        caught_n = len(e["by_model"])
        max_amt = max(e["by_model"].values(), default=0)
        return (-caught_n, -max_amt)

    sorted_entries = sorted(by_key.values(), key=_sort_key)
    row = header_row + 1
    n_fixed = len(fixed_headers)
    for e in sorted_entries:
        ws.cell(row=row, column=1, value=e["faktura"])
        ws.cell(row=row, column=2, value=e["varenummer"])
        ws.cell(row=row, column=3, value=(e["beskrivelse"] or "")[:120])
        ws.cell(row=row, column=4, value=e["antal"])
        c_unit = ws.cell(row=row, column=5, value=e["unit_price"])
        c_agreed = ws.cell(row=row, column=6, value=e["agreed_unit_price"])
        c_unit.number_format = '#,##0.00 "kr."'
        c_agreed.number_format = '#,##0.00 "kr."'
        for i, m in enumerate(model_order):
            v = e["by_model"].get(m)
            cell = ws.cell(row=row, column=n_fixed + 1 + i)
            if v is None:
                cell.value = "—"
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = round(v, 2)
                cell.number_format = '#,##0.00 "kr."'
        ws.cell(
            row=row,
            column=n_fixed + 1 + n_models,
            value=f"{len(e['by_model'])}/{n_models}",
        ).alignment = Alignment(horizontal="center")
        row += 1

    # Auto-filter + freeze
    if sorted_entries:
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{row - 1}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=4).coordinate

    # Column widths
    fixed_widths = {1: 16, 2: 16, 3: 36, 4: 8, 5: 16, 6: 18}
    for idx, w in fixed_widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w
    for i, _ in enumerate(model_order):
        ws.column_dimensions[get_column_letter(n_fixed + 1 + i)].width = 18
    ws.column_dimensions[get_column_letter(n_fixed + 1 + n_models)].width = 12


def _coerce_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _fmt_num(v) -> str:
    if v is None:
        return "?"
    try:
        return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return str(v)


def _format_discrepancies(items: list[dict]) -> str:
    """One human-readable line per discrepant line item."""
    lines = []
    for i, li in enumerate(items, start=1):
        desc = (li.get("description") or "(ukendt)").strip()
        qty = li.get("quantity")
        unit = li.get("unit_price")
        agreed = li.get("agreed_unit_price")
        diff_amt = li.get("discrepancy_amount")
        sign = ""
        if isinstance(diff_amt, (int, float)):
            sign = "supplier overopkrævet" if diff_amt > 0 else "supplier underopkrævet" if diff_amt < 0 else "ingen difference"
        qty_str = _fmt_num(qty) if qty is not None else "?"
        lines.append(
            f"{i}. {desc[:80]}\n"
            f"   Aftalt: {_fmt_num(agreed)}/stk · Faktureret: {_fmt_num(unit)}/stk · "
            f"Mængde: {qty_str}\n"
            f"   Difference: {_fmt_num(diff_amt)} DKK ({sign})"
        )
    return "\n\n".join(lines)


def _format_rebate(pred: dict) -> str:
    applied = pred.get("rebate_applied")
    expected = pred.get("expected_rebate")
    if applied is None and expected is None:
        return "(intet rabat-felt udfyldt)"
    if expected is None:
        return f"Faktureret rabat: {_fmt_num(applied)} DKK (ingen forventet rabat beregnet)"
    if applied is None:
        return f"Forventet rabat: {_fmt_num(expected)} DKK (intet rabat-felt på faktura)"
    try:
        diff = float(expected) - float(applied)
    except (TypeError, ValueError):
        diff = None
    base = (
        f"Forventet: {_fmt_num(expected)} DKK\n"
        f"Faktureret: {_fmt_num(applied)} DKK"
    )
    if diff is not None:
        if abs(diff) < 1.0:
            base += f"\nDifference: {_fmt_num(diff)} DKK (OK, indenfor tolerance)"
        else:
            base += f"\nDifference: {_fmt_num(diff)} DKK ⚠ rabat-afvigelse"
    return base


def _format_credit_note(pred: dict) -> str:
    cnh = pred.get("credit_note_handling")
    doctype = pred.get("document_type")
    if doctype == "credit_note":
        if not isinstance(cnh, dict):
            return "Kreditnota — men credit_note_handling mangler ⚠"
        ref = cnh.get("references_invoice") or "(ingen reference fundet)"
        sign = cnh.get("sign_convention") or "(ukendt)"
        return f"Kreditnota\nFortegn: {sign}\nRef. faktura: {ref}"
    if isinstance(cnh, dict) and cnh.get("is_credit_note"):
        return "Inkonsistens: document_type=invoice men credit_note_handling.is_credit_note=true ⚠"
    return "(ikke en kreditnota)"


def json_diff(gt: dict, pred: dict) -> str:
    """Compact per-field diff. Only shows keys where GT != pred."""
    lines = []
    keys = set((gt or {}).keys()) | set((pred or {}).keys())
    for k in sorted(keys):
        g = (gt or {}).get(k)
        p = (pred or {}).get(k)
        if g != p:
            lines.append(f"{k}:\n  gt:   {json.dumps(g, ensure_ascii=False, default=str)}\n  pred: {json.dumps(p, ensure_ascii=False, default=str)}")
    return "\n".join(lines) if lines else "(match)"
